import re
from datetime import datetime
from zipfile import BadZipFile

from flask import (current_app, flash, jsonify, redirect, render_template,
                   request, url_for)
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import func
from sqlalchemy.exc import SQLAlchemyError, IntegrityError

from ....modals import db
from ....modals.branches_db import Branch, BranchClasses
from ....modals.students_db import Student, StudentSubjectAllocation
from .. import admin_bp
from ..forms import AddStudentForm, StudentSearchForm
from ..forms.branches_forms import BranchesList, BranchGradeStreamForm
from ..forms.students_forms import (MuiltapleStudentsUploadForm,
                                    PassportUploadForm)
from ..utils import load_branch_choices, preprocess_image, safe_date, validate_fullname
from ..utils.route_protect import admin_required
from flask_login import login_required 

from ..services.subs import auto_allocate_subjects 

from ..services.studs import get_next_adm_no
from flask_login import current_user

@admin_bp.route("/student_dash", methods=["GET", "POST"])
@login_required
@admin_required
def student_dash(): 

    # ----------- Define Route Forms ----------- #
    form = BranchGradeStreamForm()
    add_student_form = AddStudentForm()
    muiltable_students_upload_form = MuiltapleStudentsUploadForm()
    
    add_student_form.branches.choices = load_branch_choices()

    if not current_user.is_super_admin:
        add_student_form.admission_number.data = get_next_adm_no(current_user.branch_id)

    muiltable_students_upload_form.branches.choices = load_branch_choices()

    students = []
    selected_branch = selected_grade = selected_stream = None

    branches = Branch.query.order_by(Branch.created_at.desc()).all()

    # On POST, populate form choices for validation
    if request.method == "POST":
        # Populate branch choices for WTForms validation
        form.branch.choices = [("", "--- Select Branch ---")] + [
            (str(b.id), b.branch_name) for b in branches
        ]

        branch_id = request.form.get("branch")
        grade_id = request.form.get("grade_form")

        # Populate grade choices based on selected branch
        if branch_id:
            classes = BranchClasses.query.filter_by(branch_id=branch_id).all()
            form.grade_form.choices = [("", "--- Select Grade/Form ---")] + [
                (str(c.id), c.grade_form) for c in classes
            ]

        # Populate stream choices based on selected grade/class
        if grade_id:
            class_obj = BranchClasses.query.get(grade_id)
            if class_obj and class_obj.streams:
                form.stream.choices = [("", "--- Select Stream ---")] + [
                    (s, s) for s in class_obj.streams
                ]
            else:
                form.stream.choices = []

    # Validate form and fetch students
    if form.validate_on_submit():
        branch_id = form.branch.data
        grade_id = form.grade_form.data
        stream = form.stream.data.strip() if form.stream.data else None  # optional

        # Preserve selected objects for template
        selected_branch = Branch.query.get(branch_id)
        selected_grade = BranchClasses.query.get(grade_id)
        selected_stream = stream

        # Build query
        query = Student.query.filter_by(branch_id=branch_id, class_id=grade_id)
        if selected_stream:
            query = query.filter_by(stream=selected_stream)

        students = query.order_by(Student.fullname.asc()).all()
       
    
    # Default loaded data 
    if request.method == "GET":
        students = Student.query.filter_by(branch_id=1, class_id=1).all()
        selected_branch = Branch.query.get(1)
        selected_grade = BranchClasses.query.get(1)
        
    return render_template(
        "student_templates/student_dash.html",
        select_branch_form=form,
        student_data_form=add_student_form,
        students=students,
        selected_branch=selected_branch,
        selected_grade=selected_grade,
        muiltable_students_upload_form=muiltable_students_upload_form,
        selected_stream=selected_stream,
    )


@admin_bp.route("/get_next_admission_no/<int:branch_id>")
@login_required
@admin_required
def get_next_admission_no(branch_id):
    """
    Returns next admission number for a branch.
    """

    # super admin can query any branch
    if current_user.is_super_admin:
        target_branch_id = branch_id

    # normal admin locked to own branch
    elif current_user.is_admin:
        target_branch_id = current_user.branch_id

    else:
        return jsonify({"error": "forbidden"}), 403

    next_no = get_next_adm_no(target_branch_id)
    return jsonify({"admission_no": next_no})
        

@admin_bp.route("/muiltaple_students_upload", methods=["POST"])
@login_required
@admin_required
def muiltaple_students_upload():
    form = MuiltapleStudentsUploadForm()
    form.branches.choices = load_branch_choices()

    if form.validate_on_submit():
        file = form.excel_file.data
        branch_id = form.branches.data
        grade_form_id = request.form.get("grade_form")
        stream = request.form.get("stream")

        if not grade_form_id:
            flash("Please select a grade/form to upload student data!", "danger")
            return redirect(url_for("admin.student_dash"))

        # Try reading workbook safely
        try:
            wb = load_workbook(file)  
        except OSError:
            flash("The uploaded excel has more than one workbook with data.", "danger")
            return redirect(url_for("admin.student_dash")) 
               
        except (BadZipFile, InvalidFileException):
            flash("The uploaded Excel file is invalid or corrupt.", "danger")
            return redirect(url_for("admin.student_dash"))
        
        except:
            flash("Something went wrong please try again later.", "danger")
            return redirect(url_for("admin.student_dash"))

        sheet = wb.active
        
        # Check if the uploaded file in empty
        data_rows = [
                row for row in sheet.iter_rows(min_row=2, values_only=True)
                if any(cell is not None and str(cell).strip() for cell in row)
            ]

        if not data_rows:
            flash("The uploaded Excel file is EMPTY.", "warning")
            return redirect(url_for("admin.student_dash"))

        # --------- READ HEADER ROW ----------
        headers = [
            str(h).strip().lower() if h else ""
            for h in next(sheet.iter_rows(values_only=True))
        ]

        # REQUIRED COLUMNS CHECK
        if headers[0].lower() not in [
            "adm no",
            "admission number",
            "adm_number",
            "adm",
            "admission no",
        ]:
            flash("First column must be the Admission Number.", "danger")
            return redirect(url_for("admin.student_dash"))

        if headers[1].lower() not in ["name", "fullname", "full name", "student name"]:
            flash("Second column must be the Student Name.", "danger")
            return redirect(url_for("admin.student_dash"))

        # MAP OPTIONAL FIELDS (clean + lowercase only)
        optional_map = {
            "gender": ["gender", "gendar", "sex"],
            "knec_assessment_no": [
                "assessment no",
                "assessment number",
                "assessment_no",
            ],
            "nemis_number": ["nemis", "nemis no", "nemis number"],
            "birth_cert_no": ["birth cert no", "birth certificate", "birth_cert_no"],
            "parent_fullname": [
                "parent name",
                "parent_fullname",
                "parent",
                "parent fullname",
            ],
            "parent_phone": ["parent phone", "phone", "parent_phone", "parent number"],
            "kcpe_marks": ["kcpe marks", "kcpe score", "kcpe"],
            "kcpe_year": ["kcpe year"],
            "kcpe_index_no": ["kcpe index no"],
        }

        # Create column index → database field mapping
        column_mapping = {}
        for db_field, excel_variants in optional_map.items():
            clean_variants = [v.lower() for v in excel_variants]
            for idx, col_header in enumerate(headers):
                if col_header in clean_variants:
                    column_mapping[idx] = db_field
                    break

        # Preload existing ADM numbers
        existing_adm_numbers = {
            str(s.admission_number)
            for s in Student.query.filter_by(branch_id=branch_id).all()
        }

        added = duplicate = skipped = 0
         
        for row in sheet.iter_rows(values_only=True, min_row=2):
            adm_no = str(row[0]).strip() if row[0] else ""
            fullname = str(row[1]).strip() if row[1] else ""

            # Check if adm no is digits only
            if not adm_no.isdigit():
                skipped += 1
                continue
            
            # Skip invalid rows
            if not adm_no or not fullname:
                skipped += 1 
                continue
            
            # Skip names with invalid characters
            if not validate_fullname(fullname):
                skipped += 1  
                continue

            # Skip duplicate ADM
            if adm_no in existing_adm_numbers:
                duplicate += 1
                continue

            student = Student(
                branch_id=branch_id,
                class_id=grade_form_id,
                stream=stream,
                admission_number=adm_no,
                fullname=fullname,
            )

            # Apply optional fields
            for col_idx, db_field in column_mapping.items():
                value = row[col_idx]

                # Type safety
                if db_field in ["kcpe_marks", "kcpe_year"] and value:
                    try:
                        value = int(value)
                    except:
                        value = None

                setattr(student, db_field, value)

            db.session.add(student)
            db.session.flush()  # get student.id immediately

            auto_allocate_subjects(student)

            existing_adm_numbers.add(adm_no)
            added += 1

        # Commit safely
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Database error: {e}", "danger")
            return redirect(url_for("admin.student_dash"))
        
        added_message = (
            f"<span class='text-success'>{ added } => Students added successfully</span>"
            if added > 0
            else "No student added"
        )

        skipped_message = (
            f"""<span class='text-danger'>{ skipped } => Invalid rows skipped! 
                (Empty or invalid names, invalid adm no etc. check your data and try again.)"""
            if skipped > 0
            else ""
        )
        flash(
            f"""
            <strong>Excel Upload Summary:</strong><br>
            {added_message}<br>
            {duplicate} ⇒ Duplicate student(s) skipped.<br>
            {skipped_message}
            """,
            "success" if added > 0 else "warning",
        )

        return redirect(url_for("admin.student_dash"))

    # WTForms validation errors
    for field, errors in form.errors.items():
        for error in errors:
            flash(f"{getattr(form, field).label.text}: {error}", "danger")

    return redirect(url_for("admin.student_dash"))


@admin_bp.route("/add_student", methods=["POST"])
@login_required
@admin_required
def add_student():
    form = AddStudentForm(request.form)

    if not current_user.is_super_admin:
        target_branch_id = current_user.branch_id
        form.admission_number.data = get_next_adm_no(target_branch_id)
    
    # Ensure form is being submitted
    if not form.is_submitted():
        flash("Invalid submission.", "danger")
        return redirect(url_for("admin.student_dash"))

    if form.fullname.data == "":
        flash("Student's fullname is required.", "danger")
        return redirect(url_for("admin.student_dash"))
    
    if not validate_fullname(form.fullname.data):
        flash("The student name contain invalid characters!", 'danger')
        return redirect(url_for("admin.student_dash"))
    
    # Optional WTForms validation
    form.validate()

    # ---- Extract dynamic fields ----
    class_id = request.form.get("grade_form")
    
    if request.form.get("stream") == "":
        flash("You failed to assign the student a stream.", "danger")
        return redirect(url_for("admin.student_dash"))
    
    stream = request.form.get("stream", None)


    if not class_id:
        flash("Please select a branch and a grade for the student.", "danger")
        return redirect(url_for("admin.student_dash"))

    branch_id = form.branches.data
    admission_no =  form.admission_number.data

    # ---- UNIQUE CHECK per branch ----
    existing_student = Student.query.filter_by(
        branch_id=branch_id, admission_number=admission_no
    ).first()

    if existing_student:
        flash(
            f"Admission number '{admission_no}' already exists in this branch.",
            "danger",
        )
        return redirect(url_for("admin.student_dash"))

    # ---- Create student object ----
    student = Student(
        branch_id=branch_id,
        class_id=class_id,
        stream=stream,
        admission_number=admission_no,
        fullname=form.fullname.data,
        knec_assessment_no=form.knec_assessment_no.data,
        nemis_number=form.nemis_number.data,
        birth_cert_no=form.birth_cert_no.data,
        gender=form.gender.data,
        dob=form.dob.data,
        boarding_status=form.boarding_status.data,
        pathway=form.pathway.data,
        kcpe_marks=form.kcpe_marks.data,
        kcpe_index_no=form.kcpe_index_no.data,
        kcpe_year=form.kcpe_year.data,
        date_of_admission=form.date_of_admission.data,
        parent_fullname=form.parent_fullname.data,
        parent_phone=form.parent_phone.data,
    )

    try:
        db.session.add(student)
        db.session.flush()  # 🔥 ensures student.id is available

        auto_allocate_subjects(student)

        db.session.commit()
        flash("Student added successfully!", "success")
        
    except IntegrityError as e:
        current_app.logger.exception(
            f"YOUR FORM IS MISSING SOME FIELDS: {e}"
        )
        db.session.rollback()
        flash("A student must be assigned an admission no.", "danger")
        
    except Exception as e:
        db.session.rollback()
        flash(f"Database Error: {e}", "danger")

    return redirect(url_for("admin.student_dash"))


@admin_bp.route("/student_profile/<int:student_id>", methods=["GET", "POST"])
@login_required
@admin_required
def student_profile(student_id):
    student = Student.query.get_or_404(student_id)
    student_branch = Branch.query.get(student.branch_id)
    student_class = student.class_info

    transfer_form = BranchesList()
    transfer_form.branches.choices = load_branch_choices()
    

    # Initialize WTForm for passport upload
    passport_form = PassportUploadForm() 
    if request.method == "POST":
        if passport_form.validate_on_submit():
            file = passport_form.passport.data
            student.passport = preprocess_image(file)
            try:
                db.session.commit()
            except:
                flash(
                    "Sorry you image can't be saved right now! Please tell later.",
                    "danger",
                )

            flash("Passport uploaded successfully.", "success")
            return redirect(url_for("admin.student_profile", student_id=student.id))
        else:
            flash("Failed to upload passport. Please check the errors.", "danger")

    data = {
        "id": student.id,
        "branch_name": student_branch.branch_name if student_branch else "---",
        "grade_form": student_class.grade_form if student_class else "---",
        "stream": student.stream or "",
        "admission_number": student.admission_number or "---",
        "fullname": student.fullname or "---",
        "knec_assessment_no": student.knec_assessment_no or "",
        "nemis_number": student.nemis_number or "",
        "birth_cert_no": student.birth_cert_no or "",
        "gender": student.gender or "",
        "dob": safe_date(student.dob),
        "boarding_status": student.boarding_status or "",
        "pathway": student.pathway or "",
        "kcpe_marks": student.kcpe_marks or "",
        "kcpe_index_no": student.kcpe_index_no or "",
        "kcpe_year": student.kcpe_year or "",
        "date_of_admission": safe_date(student.date_of_admission),
        "parent_fullname": student.parent_fullname or "",
        "parent_phone": student.parent_phone or "",
        "passport": student.passport, 
        "subjects_taken": student.subjects_taken
    }

    return render_template(
        "student_templates/student_profile.html",
        student=data,
        student_branch=student_branch,
        student_class=student_class,
        passport_form=passport_form,
        transfer_form=transfer_form,
    )


@admin_bp.route("/update_student/<int:student_id>", methods=["POST"])
@login_required
@admin_required
def update_student(student_id):
    student = Student.query.get_or_404(student_id)

    name = request.form.get("fullname")
    # Validate fullname
    if not validate_fullname(name):
        flash(f"Student name {name} contain invalid characters!", "danger")
        return redirect(url_for("admin.student_profile", student_id=student_id))
    try:
        # Get form data
        fullname = name
        gender = request.form.get("gender")
        dob_str = request.form.get("dob")  # <-- comes as string from form
        boarding_status = request.form.get("boarding_status")
        pathway = request.form.get("pathway")
        parent_fullname = request.form.get("parent_fullname")
        parent_phone = request.form.get("parent_phone")
        knec_assessment_no = request.form.get("knec_assessment_no")
        nemis_number = request.form.get("nemis_number")
        birth_cert_no = request.form.get("birth_cert_no")
        kcpe_marks = request.form.get("kcpe_marks")
        kcpe_index_no = request.form.get("kcpe_index_no")
        kcpe_year = request.form.get("kcpe_year")

        # Convert DOB string to date
        if dob_str:
            try:
                dob = datetime.strptime(dob_str, "%Y-%m-%d").date()
            except ValueError:
                flash("Invalid date format for Date of Birth.", "warning")
                return redirect(request.referrer)
        else:
            dob = None

        # Update the student
        student.fullname = fullname
        student.gender = gender
        student.dob = dob
        student.boarding_status = boarding_status
        student.pathway = pathway
        student.parent_fullname = parent_fullname
        student.parent_phone = parent_phone
        student.knec_assessment_no = knec_assessment_no
        student.nemis_number = nemis_number
        student.birth_cert_no = birth_cert_no
        student.kcpe_marks = kcpe_marks
        student.kcpe_index_no = kcpe_index_no
        student.kcpe_year = kcpe_year

        db.session.commit()
        flash("Student updated successfully.", "success")
    except Exception as e:
        current_app.logger.error(f"Error updating student {student_id}: {e}")
        flash("An error occurred while updating the student.", "danger")

    return redirect(url_for("admin.student_profile", student_id=student_id))


@admin_bp.route("/delete_student/<int:student_id>", methods=["POST"])
@login_required
@admin_required
def delete_student(student_id):
    try:
        student = Student.query.get(student_id)
        if not student:
            flash("Student not found.", "warning")
            return redirect(url_for("admin.student_dash"))

        db.session.delete(student)
        db.session.commit()
        flash(f"Student {student.fullname} deleted successfully.", "success")

    except SQLAlchemyError as e:
        db.session.rollback()  # undo any partial changes
        current_app.logger.error(
            f"Database error while deleting student {student_id}: {e}", exc_info=True
        )
        flash(
            "An error occurred while deleting the student. Please try again.", "danger"
        )

    except Exception as e:
        current_app.logger.error(
            f"Unexpected error while deleting student {student_id}: {e}", exc_info=True
        )
        flash("An unexpected error occurred. Please contact support.", "danger")

    return redirect(url_for("admin.student_dash"))


def serialize_students(student_list):
    return [
        {
            "id": s.id,
            "fullname": s.fullname,
            "admission_number": s.admission_number,
            "assessment_number": s.knec_assessment_no,
            "branch": s.branch.branch_name,
            "grade_form": s.class_info.grade_form,
            "stream": s.stream,
        }
        for s in student_list
    ]


@admin_bp.route("/fetch_searched_student", methods=["POST"])
@login_required
@admin_required
def fetch_searched_student():
    form = StudentSearchForm()

    if not form.validate_on_submit():
        return jsonify({"status": "error", "message": "Invalid search"}), 400

    raw_query = form.query.data.strip().lower()

    # ---------------------------------------
    # 1. ADMISSION NUMBER EXACT MATCH
    # ---------------------------------------
    if raw_query.isdigit():
        students = Student.query.filter_by(admission_number=int(raw_query)).all()
        return jsonify({"status": "success", "students": serialize_students(students)})

    # ---------------------------------------
    # 2. ASSESSMENT NUMBER EXACT MATCH
    # ---------------------------------------
    students = Student.query.filter(
        func.lower(Student.knec_assessment_no) == raw_query
    ).all()
    if students:
        return jsonify({"status": "success", "students": serialize_students(students)})

    # ---------------------------------------
    # 3. FULL NAME MATCH (ANY ORDER, WHOLE WORDS)
    # ---------------------------------------
    # remove punctuation
    normalized = re.sub(r"[^\w\s]", " ", raw_query)
    name_parts = normalized.split()

    # We build a fullname with spaces around it: " John Doe "
    fullname_spaced = func.lower(" " + Student.fullname + " ")

    query = Student.query

    for word in name_parts:
        query = query.filter(fullname_spaced.like(f"% {word} %"))

    students = query.all()

    return jsonify({"status": "success", "students": serialize_students(students)})


@admin_bp.route("/move_student/<int:student_id>", methods=["POST", "GET"])
@login_required
@admin_required
def move_student(student_id):
    transfer_form = BranchesList()
    transfer_form.branches.choices = load_branch_choices()

    try:
        student = Student.query.get(student_id)
        if not student:
            flash("Student not found.", "warning")
            return redirect(url_for("admin.student_dash"))

        # Values from form
        new_branch_id = request.form.get("branches", type=int)
        new_class_id = request.form.get("grade_form", type=int)
        new_stream = request.form.get("stream")  # Optional

        # Validate required fields
        if not new_branch_id or not new_class_id:
            flash(
                "Invalid movement request. Missing branch or class details.", "warning"
            )
            return redirect(url_for("admin.student_profile", student_id=student.id))

        # Current student data
        current_branch_id = student.branch_id
        current_class_id = student.class_id
        current_stream = student.stream or None

        # Prevent moving to SAME branch + SAME class + SAME stream
        if (
            new_branch_id == current_branch_id
            and new_class_id == current_class_id
            and (
                new_stream == current_stream or (not new_stream and not current_stream)
            )
        ):
            flash(
                "The student is already in the selected branch, grade/form, and stream.",
                "warning",
            )
            return redirect(url_for("admin.student_profile", student_id=student.id))

        # --- Only generate a new admission number if the branch is changing ---
        if new_branch_id != current_branch_id:
            largest_adm = (
                db.session.query(db.func.max(Student.admission_number))
                .filter(Student.branch_id == new_branch_id)
                .scalar()
            )
            student.admission_number = (largest_adm or 0) + 1
        # Else keep the current admission_number intact

        # Apply updates
        student.branch_id = new_branch_id
        student.class_id = new_class_id
        student.stream = new_stream or None

        db.session.commit()

        flash("Student successfully moved.", "success")
        return redirect(url_for("admin.student_profile", student_id=student.id))

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(e)
        flash("An unexpected error occurred while moving the student.", "danger")
        return redirect(url_for("admin.student_dash"))


@admin_bp.route("/students/by-class-subject", methods=["POST"])
@login_required
def students_by_class_subject():
    data = request.get_json()

    branch_id = data.get("branch_id")
    grade_form = data.get("grade_form")
    subject_id = data.get("subject_id")

    if not branch_id or not grade_form or not subject_id:
        return jsonify({"error": "Missing required data"}), 400

    class_obj = BranchClasses.query.filter_by(
        branch_id=branch_id,
        grade_form=grade_form
    ).first()

    if not class_obj:
        return jsonify([])

    students = Student.query.filter_by(
        branch_id=branch_id,
        class_id=class_obj.id
    ).order_by(Student.fullname.asc()).all()

    students_data = [
        {
            "id": s.id,
            "fullname": s.fullname,
            "admission_number": s.admission_number,
            "allocated": any(alloc.subject_id == int(subject_id) for alloc in s.subject_allocations)
        }
        for s in students
    ]
    
    allocated_count = sum(1 for s in students_data if s["allocated"])

    return jsonify(
        {
            "students": students_data,
            "allocated_count": allocated_count
        }
    )

 
@admin_bp.route("/subjects/allocate", methods=["POST"])
@login_required
def allocate_subjects():
    """
    Allocate or remove a subject for selected students.
    Expects JSON payload:
    {
        "branch_id": 1,
        "grade_form": "Grade 7",
        "subject_id": 5,
        "students": [1, 2, 3]  # students currently checked
    }
    """
    data = request.get_json()

    branch_id = data.get("branch_id")
    grade_form = data.get("grade_form")
    subject_id = data.get("subject_id")
    student_ids_checked = set(data.get("students", []))  # currently checked students

    # Basic validation
    if not branch_id or not grade_form or not subject_id:
        return jsonify({"error": "Missing required data"}), 400

    # Fetch all students of this class and branch that have this subject allocated
    existing_allocations = StudentSubjectAllocation.query\
        .join(Student)\
        .filter(Student.branch_id == branch_id)\
        .filter(Student.class_info.has(grade_form=grade_form))\
        .filter(StudentSubjectAllocation.subject_id == subject_id)\
        .all()

    already_allocated_ids = set([alloc.student_id for alloc in existing_allocations])

    # --- Students to ADD allocation ---
    to_add_ids = student_ids_checked - already_allocated_ids
    new_allocations = [StudentSubjectAllocation(student_id=sid, subject_id=subject_id) for sid in to_add_ids]

    # --- Students to REMOVE allocation ---
    to_remove_ids = already_allocated_ids - student_ids_checked
    if to_remove_ids:
        StudentSubjectAllocation.query.filter(
            StudentSubjectAllocation.subject_id == subject_id,
            StudentSubjectAllocation.student_id.in_(to_remove_ids)
        ).delete(synchronize_session=False)

    # Commit all changes
    if new_allocations or to_remove_ids:
        if new_allocations:
            db.session.add_all(new_allocations)
        db.session.commit()

    return jsonify({
        "status": "success",
        "added_count": len(new_allocations),
        "removed_count": len(to_remove_ids),
        "already_allocated_count": len(already_allocated_ids & student_ids_checked)
    })

