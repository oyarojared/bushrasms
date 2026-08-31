from ....modals.assessment_db import *
from ....modals import db
from ....modals.branches_db import Branch, BranchClasses
from ....modals.staff_db import ClassTeacher, Teacher
from ....modals.students_db import Student
from ....modals.subjects_db import *
from ..utils import resolve_grade
from ....modals.students_db import StudentSubjectAllocation

from pathlib import Path
import base64
import io
import os
from flask import current_app
from ..services.grading import get_max_points_for_class
from .grades import live_class_name
from PIL import Image

_pdf_image_cache = {}


def build_pdf_image_data_uri(source, max_size=96, quality=72):
    """
    Return a compact JPEG data URI for PDF rendering.
    WeasyPrint is much faster with embedded thumbnails than repeated file:// loads.
    """
    if not source:
        return None

    cache_key = (source, max_size, quality)
    if cache_key in _pdf_image_cache:
        return _pdf_image_cache[cache_key]

    if source.startswith("data:image"):
        _pdf_image_cache[cache_key] = source
        return source

    if source.startswith("file:///"):
        path = Path(source[8:])
    elif source.startswith("file://"):
        path = Path(source[7:])
    else:
        path = Path(source)

    if not path.exists():
        _pdf_image_cache[cache_key] = source
        return source

    try:
        image = Image.open(path).convert("RGB")
        image.thumbnail((max_size, max_size))
        buffer = io.BytesIO()
        image.save(buffer, format="JPEG", quality=quality, optimize=True)
        data_uri = "data:image/jpeg;base64," + base64.b64encode(buffer.getvalue()).decode("ascii")
        _pdf_image_cache[cache_key] = data_uri
        return data_uri
    except Exception:
        _pdf_image_cache[cache_key] = source
        return source


def build_static_image_path(filename, folder="uploads/passports", default="default-logo.PNG"):
    """
    Returns a file URI for WeasyPrint to use.
    Works with root-level uploads directory.
    """
    # Move to /home/Bushraschools
    project_root = Path(current_app.root_path).parents[2]

    base = project_root / folder

    if filename:
        clean_name = filename.strip()
        path = base / clean_name

        if path.exists():
            return path.resolve().as_uri()

    # fallback
    return (base / default).resolve().as_uri()


def build_passport_path(student):
    # Move up to /home/Bushraschools
    base = Path(current_app.root_path).parents[2] / "uploads" / "passports"

    filename = student.passport.strip() if student.passport else None

    if filename:
        path = base / filename
        if path.exists():
            return path.resolve().as_uri()

    return (base / "default.jpg").resolve().as_uri()



# def get_report_card_data(branch_id, class_id, exam_id, stream=None, student_id=None):
#     """
#     Fetch all necessary data to generate a report card PDF,
#     including grading reference for the class.
#     """

#     # Branch info
#     branch = Branch.query.get(branch_id)
#     if not branch:
#         raise ValueError("Branch not found")

#     # Class info
#     class_ = BranchClasses.query.get(class_id)
#     if not class_:
#         raise ValueError("Class not found")

#     class_name = class_.grade_form

#     # Exam info
#     exam_data = Exam.query.get(exam_id)

#     # Class teacher
#     class_teacher_query = ClassTeacher.query.filter_by(
#         branch_id=branch_id,
#         class_id=class_id
#     )

#     if stream:
#         class_teacher_query = class_teacher_query.filter_by(stream=stream)

#     class_teacher_obj = class_teacher_query.first()
#     class_teacher_name = (
#         class_teacher_obj.teacher.fullname
#         if class_teacher_obj and class_teacher_obj.teacher
#         else None
#     )

#     # =========================
#     # FETCH ALL STUDENTS FIRST
#     # =========================
#     students = Student.query.filter_by(
#         branch_id=branch_id,
#         class_id=class_id
#     ).all()

#     student_list = []

#     # =========================
#     # BUILD STUDENT DATA
#     # =========================
#     for s in students:

#         student_data = {
#             "id": s.id,
#             "fullname": s.fullname.upper(),
#             "assessment_no": s.knec_assessment_no,
#             "admission_number": s.admission_number,
#             "pathway": s.pathway,
#             "gender": s.gender,
#             "stream": s.stream,
#             "passport_path": build_passport_path(s),
#             "class_teacher": class_teacher_name,
#             "subjects": []
#         }

#         for alloc in s.subject_allocations:
#             subject = alloc.subject
#             if not subject:
#                 continue

#             lesson = Lesson.query.filter_by(
#                 branch_id=branch_id,
#                 class_id=class_id,
#                 stream=stream,
#                 subject_id=subject.id
#             ).first()

#             teacher_initials = None
#             if lesson and lesson.teacher:
#                 names = lesson.teacher.fullname.strip().split()
#                 teacher_initials = ".".join([n[0].upper() for n in names])

#             exam_paper = ExamPaper.query.filter_by(
#                 exam_id=exam_id,
#                 branch_id=branch_id,
#                 class_id=class_id,
#                 stream=stream,
#                 subject_id=subject.id
#             ).first()

#             marks = None

#             if exam_paper:
#                 mark_obj = StudentExamMark.query.filter_by(
#                     exam_paper_id=exam_paper.id,
#                     student_id=s.id
#                 ).first()

#                 if mark_obj:
#                     marks = mark_obj.marks

#                     if marks:
#                         student_data["total_marks"] += marks

#             grade_info = resolve_grade(class_id, marks) if marks is not None else {
#                 "performance_level": None,
#                 "points": None,
#                 "descriptor": None
#             }

#             student_data["subjects"].append({
#                 "subject_code": subject.code,
#                 "subject_name": subject.name,
#                 "teacher_initials": teacher_initials,
#                 "marks": marks,
#                 "performance_level": grade_info["performance_level"],
#                 "points": grade_info["points"],
#                 "descriptor": grade_info["descriptor"]
#             })

#         student_list.append(student_data)

#     # ======================================================
#     # STEP 1: CLASS RANKING (FULL CLASS - NOT FILTERED)
#     # ======================================================
#     student_list.sort(
#         key=lambda student: student["total_marks"],
#         reverse=True
#     )

#     for position, student in enumerate(student_list, start=1):
#         student["class_position"] = position
#         student["class_total"] = len(student_list)

#     # ======================================================
#     # STEP 2: STREAM RANKING (FULL CLASS)
#     # ======================================================
#     has_streams = any(s.get("stream") for s in student_list)

#     if has_streams:
#         from collections import defaultdict

#         stream_groups = defaultdict(list)

#         for student in student_list:
#             stream_key = student.get("stream")
#             if stream_key:
#                 stream_groups[stream_key].append(student)

#         for stream_students in stream_groups.values():

#             stream_students.sort(
#                 key=lambda s: s["total_marks"],
#                 reverse=True
#             )

#             for position, student in enumerate(stream_students, start=1):
#                 student["stream_position"] = position
#                 student["stream_total"] = len(stream_students)

#     else:
#         for student in student_list:
#             student["stream_position"] = None

#     # ======================================================
#     # STEP 3: APPLY REQUEST FILTERING (DISPLAY ONLY)
#     # ======================================================
#     if stream:
#         student_list = [
#             s for s in student_list
#             if s.get("stream") == stream
#         ]

#     if student_id:
#         student_list = [
#             s for s in student_list
#             if s["id"] == student_id
#         ]

#     # Final sort for display only
#     student_list.sort(
#         key=lambda student: student["total_marks"],
#         reverse=True
#     )

#     # ======================================================
#     # Grading boundaries
#     # ======================================================
#     grading_boundaries = []
#     scheme_link = GradeGradingScheme.query.filter_by(grade_id=class_id).first()
#     if scheme_link and scheme_link.scheme:
#         grading_boundaries = scheme_link.scheme.boundaries

#     # ======================================================
#     # Logo
#     # ======================================================
#     school_logo_path = None
#     if branch.logo:
#         school_logo_path = build_static_image_path(branch.logo)

#     # ======================================================
#     # RESULT
#     # ======================================================
#     result = {
#         "branch": {
#             "name": branch.branch_name.upper(),
#             "code": branch.school_code,
#             "class_name": class_name,
#             "logo": school_logo_path,
#             "motto": branch.motto,
#         },
#         "exam": {
#             "name": exam_data.name,
#             "year": exam_data.year,
#             "term": exam_data.term
#         },
#         "class": {
#             "grade_form": class_.grade_form,
#             "class_year": class_.class_year,
#             "streams": class_.streams
#         },
#         "exam_id": exam_id,
#         "students": student_list,
#         "grading_boundaries": grading_boundaries,
#         "school_logo": None,
#         "stamp_placeholder": True,
#         "max_points": get_max_points_for_class(class_id)
#     }

#     return result




from collections import defaultdict
from sqlalchemy import func
from sqlalchemy.orm import joinedload


def _report_student_options():
    return (
        joinedload(Student.subject_allocations).joinedload(StudentSubjectAllocation.subject),
        joinedload(Student.branch),
        joinedload(Student.class_info),
    )


def student_ids_who_sat_exam(branch_id, class_id, exam_id):
    rows = (
        db.session.query(StudentExamMark.student_id)
        .join(ExamPaper, StudentExamMark.exam_paper_id == ExamPaper.id)
        .filter(
            ExamPaper.exam_id == int(exam_id),
            ExamPaper.branch_id == int(branch_id),
            ExamPaper.class_id == int(class_id),
        )
        .distinct()
        .all()
    )
    return {student_id for (student_id,) in rows}


def sitting_stream_by_student(branch_id, class_id, exam_id):
    """Stream each learner sat in for this exam, from the marksheet papers."""
    rows = (
        db.session.query(StudentExamMark.student_id, ExamPaper.stream)
        .join(ExamPaper, StudentExamMark.exam_paper_id == ExamPaper.id)
        .filter(
            ExamPaper.exam_id == int(exam_id),
            ExamPaper.branch_id == int(branch_id),
            ExamPaper.class_id == int(class_id),
        )
        .all()
    )
    streams = {}
    for student_id, paper_stream in rows:
        current = streams.get(student_id)
        if student_id not in streams:
            streams[student_id] = paper_stream
        elif current in (None, "") and paper_stream not in (None, ""):
            streams[student_id] = paper_stream
    return streams


def load_class_students_for_report(
    branch_id, class_id, include_student_id=None, exam_id=None
):
    """
    Class-wide printing uses the current class roster.

    A single learner's historical card ranks against everyone who sat
    that exam in that class, including classmates who have since moved.
    """
    query_options = _report_student_options()

    if include_student_id and exam_id:
        sat_ids = student_ids_who_sat_exam(branch_id, class_id, exam_id)
        sat_ids.add(int(include_student_id))
        return (
            Student.query
            .options(*query_options)
            .filter(Student.id.in_(sat_ids))
            .all()
        )

    students = (
        Student.query
        .options(*query_options)
        .filter_by(branch_id=branch_id, class_id=class_id)
        .all()
    )
    if not include_student_id:
        return students

    include_student_id = int(include_student_id)
    if any(student.id == include_student_id for student in students):
        return students

    extra = (
        Student.query
        .options(*query_options)
        .filter_by(id=include_student_id)
        .first()
    )
    if extra:
        students.append(extra)
    return students


def _paper_for_subject(papers_by_stream_subject, stream_key, subject_id):
    paper = papers_by_stream_subject.get((stream_key, subject_id))
    if paper:
        return paper
    for (stored_stream, stored_subject_id), stored_paper in papers_by_stream_subject.items():
        if stored_subject_id == subject_id:
            return stored_paper
    return None


def _lesson_for_subject(lessons_by_stream_subject, stream_key, subject_id):
    lesson = lessons_by_stream_subject.get((stream_key, subject_id))
    if lesson:
        return lesson
    for (stored_stream, stored_subject_id), stored_lesson in lessons_by_stream_subject.items():
        if stored_subject_id == subject_id:
            return stored_lesson
    return None


def compute_cbe_exam_rankings(branch_id, class_id, exam_id, include_student_id=None):
    """
    Lightweight class rankings for CBE exams.
    Returns a dict keyed by student_id with stream/overall positions.
    """
    students = load_class_students_for_report(
        branch_id,
        class_id,
        include_student_id=include_student_id,
        exam_id=exam_id if include_student_id else None,
    )
    sitting_streams = (
        sitting_stream_by_student(branch_id, class_id, exam_id)
        if include_student_id else {}
    )

    total_rows = (
        db.session.query(
            StudentExamMark.student_id,
            func.coalesce(func.sum(StudentExamMark.marks), 0),
        )
        .join(ExamPaper, StudentExamMark.exam_paper_id == ExamPaper.id)
        .filter(
            ExamPaper.exam_id == exam_id,
            ExamPaper.branch_id == branch_id,
            ExamPaper.class_id == class_id,
        )
        .group_by(StudentExamMark.student_id)
        .all()
    )
    totals = {student_id: float(total) for student_id, total in total_rows}

    student_rows = [
        {
            "id": student.id,
            "stream": sitting_streams.get(student.id, "")
            if include_student_id
            else student.stream,
            "total_marks": totals.get(student.id, 0.0),
        }
        for student in students
    ]

    overall_students = sorted(
        student_rows,
        key=lambda row: row["total_marks"],
        reverse=True,
    )
    class_total = len(overall_students)
    ranking_map = {}

    for position, row in enumerate(overall_students, start=1):
        ranking_map[row["id"]] = {
            "overall_position": position,
            "overall_total": class_total,
            "stream_position": None,
            "stream_total": None,
        }

    stream_groups = defaultdict(list)
    for row in overall_students:
        stream_groups[row["stream"]].append(row)

    for group in stream_groups.values():
        group.sort(key=lambda row: row["total_marks"], reverse=True)
        stream_total = len(group)
        for position, row in enumerate(group, start=1):
            ranking_map[row["id"]]["stream_position"] = position
            ranking_map[row["id"]]["stream_total"] = stream_total

    return ranking_map


def get_report_card_data(branch_id, class_id, exam_id, stream=None, student_id=None):
    """
    Fetch all necessary data to generate a report card PDF,
    including grading reference for the class.
    """

    # ------------------------------------------------------------------
    # Branch
    # ------------------------------------------------------------------
    branch = Branch.query.get(branch_id)
    if not branch:
        raise ValueError("Branch not found")

    # ------------------------------------------------------------------
    # Class
    # ------------------------------------------------------------------
    class_ = BranchClasses.query.get(class_id)
    if not class_:
        raise ValueError("Class not found")

    class_name = live_class_name(class_.grade_form)

    # ------------------------------------------------------------------
    # Exam
    # ------------------------------------------------------------------
    exam_data = Exam.query.get(exam_id)

    # ------------------------------------------------------------------
    # Class teacher
    # ------------------------------------------------------------------
    class_teacher_query = ClassTeacher.query.filter_by(
        branch_id=branch_id,
        class_id=class_id
    )

    if stream:
        class_teacher_query = class_teacher_query.filter_by(stream=stream)

    class_teacher_obj = class_teacher_query.first()

    class_teacher_name = (
        class_teacher_obj.teacher.fullname
        if class_teacher_obj and class_teacher_obj.teacher
        else None
    )

    # ------------------------------------------------------------------
    # IMPORTANT:
    # Always load ALL students in the class.
    # Ranking is calculated before filtering by stream or student.
    # ------------------------------------------------------------------
    students = load_class_students_for_report(
        branch_id,
        class_id,
        include_student_id=student_id,
        exam_id=exam_id if student_id else None,
    )
    sitting_streams = (
        sitting_stream_by_student(branch_id, class_id, exam_id)
        if student_id else {}
    )

    mark_rows = (
        db.session.query(StudentExamMark, ExamPaper, Subject)
        .join(ExamPaper, StudentExamMark.exam_paper_id == ExamPaper.id)
        .join(Subject, ExamPaper.subject_id == Subject.id)
        .filter(
            ExamPaper.exam_id == exam_id,
            ExamPaper.branch_id == branch_id,
            ExamPaper.class_id == class_id,
        )
        .all()
    )

    marks_by_student_subject = {}
    papers_by_stream_subject = {}
    for mark, paper, subject in mark_rows:
        marks_by_student_subject[(mark.student_id, subject.id)] = mark.marks
        stream_key = paper.stream if paper.stream not in (None, "") else ""
        papers_by_stream_subject[(stream_key, subject.id)] = paper

    lessons = (
        Lesson.query
        .filter_by(branch_id=branch_id, class_id=class_id)
        .options(joinedload(Lesson.teacher))
        .all()
    )
    lessons_by_stream_subject = {}
    for lesson in lessons:
        stream_key = lesson.stream if lesson.stream not in (None, "") else ""
        lessons_by_stream_subject[(stream_key, lesson.subject_id)] = lesson

    class_teachers = (
        ClassTeacher.query
        .filter_by(branch_id=branch_id, class_id=class_id)
        .options(joinedload(ClassTeacher.teacher))
        .all()
    )
    class_teachers_by_stream = {}
    for class_teacher in class_teachers:
        stream_key = class_teacher.stream if class_teacher.stream not in (None, "") else ""
        class_teachers_by_stream[stream_key] = class_teacher

    student_list = []

    # ==================================================================
    # Build student data
    # ==================================================================
    for s in students:
        sitting_stream = s.stream
        if student_id:
            sitting_stream = sitting_streams.get(s.id, "")
            if int(student_id) == s.id and stream not in (None, ""):
                sitting_stream = stream
        stream_key = sitting_stream if sitting_stream not in (None, "") else ""

        if stream is None:
            class_teacher_obj = class_teachers_by_stream.get(stream_key)
            teacher_name = (
                class_teacher_obj.teacher.fullname
                if class_teacher_obj and class_teacher_obj.teacher
                else class_teacher_name
            )
        else:
            teacher_name = class_teacher_name

        student_data = {
            "id": s.id,
            "fullname": s.fullname.upper(),
            "assessment_no": s.knec_assessment_no,
            "admission_number": s.admission_number,
            "pathway": s.pathway,
            "gender": s.gender,
            "stream": sitting_stream,
            "passport_path": build_pdf_image_data_uri(build_passport_path(s)),
            "class_teacher": teacher_name,
            "subjects": [],
            "total_marks": 0
        }

        report_subjects = []
        seen_subject_ids = set()
        for alloc in s.subject_allocations:
            subject = alloc.subject
            if subject and subject.id not in seen_subject_ids:
                report_subjects.append(subject)
                seen_subject_ids.add(subject.id)
        for student_id_key, subject_id in marks_by_student_subject:
            if student_id_key != s.id or subject_id in seen_subject_ids:
                continue
            subject = Subject.query.get(subject_id)
            if subject:
                report_subjects.append(subject)
                seen_subject_ids.add(subject_id)

        for subject in report_subjects:
            lesson = _lesson_for_subject(
                lessons_by_stream_subject, stream_key, subject.id
            )
            teacher_initials = None

            if lesson and lesson.teacher:
                names = lesson.teacher.fullname.strip().split()
                teacher_initials = ".".join(
                    n[0].upper() for n in names
                )

            marks = marks_by_student_subject.get((s.id, subject.id))
            if marks is None:
                student_data["subjects"].append({
                    "subject_code": subject.code,
                    "subject_name": subject.name,
                    "teacher_initials": teacher_initials,
                    "marks": None,
                    "performance_level": None,
                    "points": None,
                    "descriptor": None,
                })
                continue

            student_data["total_marks"] += marks

            grade_info = resolve_grade(class_id, marks)

            student_data["subjects"].append({
                "subject_code": subject.code,
                "subject_name": subject.name,
                "teacher_initials": teacher_initials,
                "marks": marks,
                "performance_level": grade_info["performance_level"],
                "points": grade_info["points"],
                "descriptor": grade_info["descriptor"]
            })

        student_list.append(student_data)

    # ==================================================================
    # OVERALL RANKING
    # ==================================================================
    overall_students = sorted(
        student_list,
        key=lambda x: x["total_marks"],
        reverse=True
    )

    class_total = len(overall_students)

    for position, student in enumerate(overall_students, start=1):
        student["overall_position"] = position
        student["class_total"] = class_total

    # ==================================================================
    # STREAM RANKING
    # ==================================================================
    stream_groups = defaultdict(list)

    for student in overall_students:
        stream_groups[student["stream"]].append(student)

    for group in stream_groups.values():
        group.sort(
            key=lambda x: x["total_marks"],
            reverse=True
        )

        stream_total = len(group)

        for position, student in enumerate(group, start=1):
            student["stream_position"] = position
            student["stream_total"] = stream_total

    # ==================================================================
    # Filter only after rankings have been computed
    # ==================================================================
    if student_id:
        output_students = [
            s for s in overall_students
            if s["id"] == int(student_id)
        ]
        if not output_students:
            raise ValueError(f"Student {student_id} not found in class rankings")
    elif stream:
        output_students = [
            s for s in overall_students
            if s["stream"] == stream
        ]
    else:
        output_students = overall_students

    # ------------------------------------------------------------------
    # Grading boundaries
    # ------------------------------------------------------------------
    grading_boundaries = []

    scheme_link = GradeGradingScheme.query.filter_by(
        grade_id=class_id
    ).first()

    if scheme_link and scheme_link.scheme:
        grading_boundaries = scheme_link.scheme.boundaries

    # ------------------------------------------------------------------
    # Logo
    # ------------------------------------------------------------------
    school_logo_path = None

    if branch.logo:
        school_logo_path = build_pdf_image_data_uri(build_static_image_path(branch.logo))

    # ------------------------------------------------------------------
    # Result
    # ------------------------------------------------------------------
    result = {
        "branch": {
            "name": branch.branch_name.upper(),
            "email": branch.email,
            "phone": "",
            "code": branch.school_code,
            "class_name": class_name,
            "logo": school_logo_path,
            "motto": branch.motto,
        },
        "exam": {
            "name": exam_data.name,
            "year": exam_data.year,
            "term": exam_data.term
        },
        "class": {
            "grade_form": live_class_name(class_.grade_form),
            "class_year": class_.class_year,
            "streams": class_.streams
        },
        "exam_id": exam_id,
        "students": output_students,
        "grading_boundaries": grading_boundaries,
        "school_logo": None,
        "stamp_placeholder": True,
        "max_points": get_max_points_for_class(class_id)
    }
 
    return result


def build_broadsheet_data(branch_id, class_id, exam_id, stream=None):
    """
    Service function that builds and returns broadsheet data.
    Returns a dict identical to what the route currently returns.
    """

    if not all([branch_id, class_id, exam_id]):
        raise ValueError("branch_id, class_id, and exam_id are required")

    if stream in ("", "null"):
        stream = None

    try:
        from .grading_844 import (
            is_844_form,
            is_low_844_grade,
            normalize_form_name,
            resolve_844_grade,
            simplify_844_grade,
        )

        # -------------------- 1. Class & Exam --------------------
        class_obj = db.session.get(BranchClasses, class_id)
        class_name = live_class_name(class_obj.grade_form) if class_obj else "N/A"
        normalized_form = normalize_form_name(class_name)
        is_844 = is_844_form(normalized_form)

        exam_obj = ExamPaper.query.filter_by(exam_id=exam_id).first()
        exam_name = exam_obj.exam.name if exam_obj and exam_obj.exam else "N/A"

        branch = db.session.get(Branch, branch_id)
        branch_name = branch.branch_name if branch else "N/A"

        # -------------------- 2. Students --------------------
        students_query = Student.query.filter_by(branch_id=branch_id, class_id=class_id)
        if stream:
            students_query = students_query.filter_by(stream=stream)

        students = students_query.order_by(Student.fullname).all()
        student_ids = [s.id for s in students]

        if not students:
            return {
                "subjects": [],
                "students": [],
                "class_name": class_name,
                "exam_name": exam_name,
                "total_learners": 0,
                "branch_name": branch_name,
                "grading_type": "844" if is_844 else "cbc",
            }

        # -------------------- 3. Subjects --------------------
        allocations = StudentSubjectAllocation.query.filter(
            StudentSubjectAllocation.student_id.in_(student_ids)
        ).all()

        allocations_by_student = {}
        for allocation in allocations:
            allocations_by_student.setdefault(allocation.student_id, set()).add(
                allocation.subject_id
            )

        subject_ids = set(a.subject_id for a in allocations)

        subjects = Subject.query.filter(Subject.id.in_(subject_ids)).all()
        subject_map = {s.id: s for s in subjects}

        # -------------------- 4. Exam Papers --------------------
        papers_query = ExamPaper.query.filter_by(
            exam_id=exam_id,
            branch_id=branch_id,
            class_id=class_id
        ).filter(ExamPaper.subject_id.in_(subject_ids))

        if stream:
            papers_query = papers_query.filter_by(stream=stream)
        else:
            papers_query = papers_query.filter(ExamPaper.stream.is_(None))

        papers = papers_query.all()
        paper_map = {p.subject_id: p for p in papers}

        # -------------------- 5. Marks --------------------
        paper_ids = [p.id for p in papers]

        marks = StudentExamMark.query.filter(
            StudentExamMark.exam_paper_id.in_(paper_ids),
            StudentExamMark.student_id.in_(student_ids)
        ).all()

        marks_map = {(m.student_id, m.exam_paper_id): m.marks for m in marks}

        # -------------------- 6. Teachers --------------------
        lessons_query = Lesson.query.filter_by(branch_id=branch_id, class_id=class_id)
        lessons_query = lessons_query.filter(Lesson.subject_id.in_(subject_ids))

        if stream:
            lessons_query = lessons_query.filter_by(stream=stream)

        lessons = lessons_query.all()
        lesson_map = {l.subject_id: l for l in lessons}

        teacher_ids = [l.teacher_id for l in lessons if l.teacher_id]
        teachers = Teacher.query.filter(Teacher.id.in_(teacher_ids)).all()
        teacher_map = {t.id: t for t in teachers}

        # -------------------- 6B. Class Teacher --------------------
        class_teacher = None
        class_teacher_obj = ClassTeacher.query.filter_by(
            class_id=class_id,
            branch_id=branch_id
        ).first()

        if class_teacher_obj:
            teacher = teacher_map.get(class_teacher_obj.teacher_id) or db.session.get(Teacher, class_teacher_obj.teacher_id)
            if teacher:
                class_teacher = f"{teacher.title} {teacher.fullname}"

        # -------------------- 7. Subjects Info --------------------
        subjects_data = []

        for s in subjects:
            lesson = lesson_map.get(s.id)
            teacher_name = "N/A"

            if lesson and lesson.teacher_id:
                teacher = teacher_map.get(lesson.teacher_id)
                if teacher:
                    teacher_name = f"{teacher.title} {teacher.fullname}"

            subjects_data.append({
                "id": s.id,
                "name": s.name,
                "code": s.code,
                "category": s.category or "",
                "teacher": teacher_name
            })

        # -------------------- 8. Build Students + Analytics --------------------
        students_data = []

        subject_analysis = {subj.id: {} for subj in subjects}
        subject_totals = {subj.id: [] for subj in subjects}
        at_risk_learners = []

        for s in students:
            marks_per_subject = {}

            for subj in subjects:
                paper = paper_map.get(subj.id)

                mark_value = "-"
                grade_value = None
                points_value = None

                if paper:
                    mark_value = marks_map.get((s.id, paper.id), "-")

                    if mark_value != "-":
                        if isinstance(mark_value, (int, float)):
                            mark_value = int(round(mark_value))

                        if is_844:
                            grade_value, points_value = resolve_844_grade(
                                mark_value,
                                subj.category or "",
                            )
                            if grade_value:
                                subject_analysis[subj.id][grade_value] = (
                                    subject_analysis[subj.id].get(grade_value, 0) + 1
                                )
                        else:
                            grade_info = resolve_grade(class_id, mark_value)
                            grade_value = (
                                grade_info.get("performance_level")
                                if grade_info
                                else None
                            )
                            points_value = (
                                grade_info.get("points") if grade_info else None
                            )
                            if grade_value:
                                subject_analysis[subj.id][grade_value] = (
                                    subject_analysis[subj.id].get(grade_value, 0) + 1
                                )

                        subject_totals[subj.id].append(mark_value)

                marks_per_subject[subj.id] = {
                    "marks": mark_value,
                    "grade": grade_value,
                    "points": points_value,
                }

            if is_844:
                low_subjects = [
                    subj.name
                    for subj in subjects
                    if is_low_844_grade(marks_per_subject[subj.id].get("grade"))
                ]
                low_count = len(low_subjects)
            else:
                low_subjects = [
                    subj.name
                    for subj in subjects
                    if marks_per_subject[subj.id].get("grade") == "BE"
                ]
                low_count = len(low_subjects)

            if low_count >= 3:
                at_risk_learners.append({
                    "id": s.id,
                    "name": s.fullname.upper(),
                    "low_subjects": ", ".join(low_subjects),
                    "low_count": low_count,
                })

            total_points, mean_grade = _broadsheet_student_summary(
                marks_per_subject,
                subjects,
                is_844,
                class_id,
            )

            students_data.append({
                "id": s.id,
                "admission_number": s.admission_number,
                "full_name": s.fullname.upper(),
                "gender": s.gender,
                "marks": marks_per_subject,
                "total_points": total_points,
                "mean_grade": mean_grade,
            })

        # -------------------- 9. Additional Analytics --------------------
        subject_averages = {}

        for subj in subjects:
            values = subject_totals[subj.id]
            subject_averages[subj.id] = (
                int(round(sum(values) / len(values))) if values else None
            )

        subject_participation = {
            subj.id: len(subject_totals[subj.id])
            for subj in subjects
        }

        # Missing marks per student (only allocated / doing subjects)
        missing_marks_list = []

        for student_obj, student_row in zip(students, students_data):
            allocated_subject_ids = allocations_by_student.get(student_obj.id, set())
            missing_subjects = []

            for subject_id in sorted(allocated_subject_ids, key=lambda sid: subject_map.get(sid).name if subject_map.get(sid) else ""):
                subject = subject_map.get(subject_id)
                if not subject:
                    continue

                mark_info = student_row["marks"].get(subject_id, {})
                if mark_info.get("marks") == "-":
                    missing_subjects.append(subject.name)

            if missing_subjects:
                missing_marks_list.append({
                    "id": student_obj.id,
                    "student": student_row["full_name"],
                    "admission_number": student_row.get("admission_number"),
                    "subjects": missing_subjects,
                })

        # -------------------- 10. Final Data --------------------
        return {
            "class_name": class_name,
            "stream": stream,
            "exam_name": exam_name,
            "subjects": subjects_data,
            "students": students_data,
            "total_learners": len(students),
            "class_teacher": class_teacher,
            "subject_analysis": subject_analysis,
            "subject_averages": subject_averages,
            "subject_participation": subject_participation,
            "at_risk_learners": at_risk_learners,
            "missing_marks": missing_marks_list,
            "branch_name": branch_name,
            "grading_type": "844" if is_844 else "cbc",
        }

    except Exception:
        current_app.logger.exception("Error building broadsheet (service)")
        raise


def _broadsheet_class_stream_label(data):
    class_name = str(data.get("class_name") or "").strip()
    stream = str(data.get("stream") or "").strip()
    if class_name and stream:
        return f"{class_name} {stream}"
    return class_name or stream or "class"


def broadsheet_excel_filename(data):
    label = _broadsheet_class_stream_label(data)
    safe = "".join(
        char if char.isalnum() or char in " ._-()" else "_"
        for char in label
    ).strip(" ._")
    return f"{safe or 'Broadsheet'}.xlsx"


def _excel_subject_label(name):
    raw = " ".join(str(name or "").split())
    if not raw:
        return "Subject"
    known = {
        "agriculture": "Agriculture",
        "arabic": "Arabic",
        "biology": "Biology",
        "business studies": "Business",
        "chemistry": "Chemistry",
        "christian religious education": "CRE",
        "computer studies": "Computer",
        "creative arts": "Creative Arts",
        "cre": "CRE",
        "english": "English",
        "environmental activities": "Environment",
        "fasihi ya kiswahili": "Fasihi",
        "french": "French",
        "geography": "Geography",
        "german": "German",
        "health education": "Health",
        "hindu religious education": "HRE",
        "history": "History",
        "history and citizenship": "History",
        "home science": "Home Science",
        "hygiene and nutrition": "Hygiene",
        "ict": "ICT",
        "indigenous language": "Indigenous",
        "indigenous languages": "Indigenous",
        "integrated science": "Int. Science",
        "integrated sciences": "Int. Science",
        "islamic religious education": "IRE",
        "kenyan sign language": "KSL",
        "kiswahili": "Kiswahili",
        "life skills": "Life Skills",
        "literature in english": "Literature",
        "mandarin": "Mandarin",
        "mathematics": "Maths",
        "maths": "Maths",
        "music": "Music",
        "physical education": "PE",
        "physics": "Physics",
        "pre technical studies": "Pre-Technical",
        "pre-technical studies": "Pre-Technical",
        "religious education": "RE",
        "science and technology": "Sci & Tech",
        "social studies": "Social Studies",
    }
    mapped = known.get(raw.lower())
    if mapped:
        return mapped
    if len(raw) <= 16:
        return raw
    shortened = raw.lower()
    for tail in (" education", " studies", " activities", " language"):
        if shortened.endswith(tail) and len(shortened) > len(tail) + 3:
            shortened = shortened[: -len(tail)].strip()
    words = [word for word in shortened.split() if word not in {"and", "&", "of", "the", "ya"}]
    label = " ".join(words).title() if words else raw
    parts = label.split()
    if len(label) > 18 and len(parts) > 2:
        return " ".join(parts[:2])
    return label


def _excel_safe(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    if text.startswith(("=", "+", "-", "@")):
        return f"'{text}"
    return text


def _broadsheet_numeric_marks(student, subjects):
    values = []
    for subject in subjects:
        mark_info = (student.get("marks") or {}).get(subject["id"], {})
        raw = mark_info.get("marks")
        if raw in (None, "-", ""):
            continue
        try:
            values.append(float(raw))
        except (TypeError, ValueError):
            continue
    return values


def build_broadsheet_excel(data, include_grades=True):
    """Spreadsheet of learners × subjects for one exam."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    subjects = data.get("subjects") or []
    students = data.get("students") or []
    include_grades = bool(include_grades)

    wb = Workbook()
    ws = wb.active
    ws.title = "Broadsheet"

    title_font = Font(bold=True, size=14, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="2C3E50")
    title_fill = PatternFill("solid", fgColor="FF7979")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")

    headers = ["Adm", "Name"]
    for subject in subjects:
        label = _excel_subject_label(subject.get("name") or subject.get("code") or "Subject")
        headers.append(label)
        if include_grades:
            headers.append(f"{label} Grade")
    headers.extend(["Total", "Mean", "Total Points", "Mean Grade"])

    last_col = get_column_letter(len(headers))
    class_stream = _broadsheet_class_stream_label(data)
    subtitle = " · ".join(
        part
        for part in [
            class_stream,
            data.get("exam_name") or "",
            "Broadsheet",
        ]
        if part
    )
    school_name = str(data.get("branch_name") or "Broadsheet").strip().upper()

    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = _excel_safe(school_name)
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A2:{last_col}2")
    sub_cell = ws["A2"]
    sub_cell.value = _excel_safe(subtitle)
    sub_cell.font = Font(bold=True, size=11, color="1F2937")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_index, student in enumerate(students, start=5):
        numeric = _broadsheet_numeric_marks(student, subjects)
        total = round(sum(numeric)) if numeric else None
        mean = round(sum(numeric) / len(numeric), 1) if numeric else None

        ws.cell(row=row_index, column=1, value=_excel_safe(student.get("admission_number"))).alignment = center
        ws.cell(row=row_index, column=2, value=_excel_safe(student.get("full_name"))).alignment = left

        col = 3
        marks = student.get("marks") or {}
        for subject in subjects:
            mark_info = marks.get(subject["id"], {})
            raw = mark_info.get("marks")
            mark_cell = ws.cell(row=row_index, column=col)
            if raw in (None, "-", ""):
                mark_cell.value = None
            else:
                try:
                    mark_cell.value = int(round(float(raw)))
                except (TypeError, ValueError):
                    mark_cell.value = _excel_safe(raw)
            mark_cell.alignment = center
            col += 1
            if include_grades:
                grade_cell = ws.cell(
                    row=row_index,
                    column=col,
                    value=_excel_safe(mark_info.get("grade") or ""),
                )
                grade_cell.alignment = center
                col += 1

        ws.cell(row=row_index, column=col, value=total).alignment = center
        ws.cell(row=row_index, column=col + 1, value=mean).alignment = center
        points = student.get("total_points")
        ws.cell(
            row=row_index,
            column=col + 2,
            value=points if points not in (None, "") else None,
        ).alignment = center
        ws.cell(
            row=row_index,
            column=col + 3,
            value=_excel_safe(student.get("mean_grade") or ""),
        ).alignment = center

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{last_col}{max(4, 4 + len(students))}"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[4].height = 22

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        values = [
            str(cell.value)
            for cell in ws[letter]
            if cell.row >= 4 and cell.value not in (None, "")
        ]
        max_len = max((len(v) for v in values), default=8)
        ws.column_dimensions[letter].width = max(8, min(max_len + 2, 28))
    ws.column_dimensions["A"].width = min(ws.column_dimensions["A"].width or 10, 12)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 18, 22)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


GRADE_ORDER_844 = [
    "A", "A-", "B+", "B", "B-", "C+", "C", "C-", "D+", "D", "D-", "E",
]
GRADE_ORDER_CBC = ["EE", "ME", "AE", "BE"]


def _is_male(gender):
    return bool(gender and str(gender).lower().startswith("m"))


def _is_female(gender):
    return bool(gender and str(gender).lower().startswith("f"))


def _grade_css_class(grade):
    if not grade:
        return ""
    return "g-" + str(grade).replace("+", "p").replace("-", "m")


def _sort_grade_counts(grade_counts, grading_type):
    order = GRADE_ORDER_844 if grading_type == "844" else GRADE_ORDER_CBC
    return [
        {
            "grade": grade,
            "count": grade_counts[grade],
            "css_class": _grade_css_class(grade),
        }
        for grade in order
        if grade_counts.get(grade, 0) > 0
    ]


def _build_subject_grade_table(subject_grade_breakdown, grading_type):
    grade_set = set()
    for block in subject_grade_breakdown:
        for item in block["grades"]:
            grade_set.add(item["grade"])

    order = GRADE_ORDER_844 if grading_type == "844" else GRADE_ORDER_CBC
    columns = [grade for grade in order if grade in grade_set]

    rows = []
    for block in subject_grade_breakdown:
        counts = {item["grade"]: item["count"] for item in block["grades"]}
        rows.append({
            "subject": block["subject"],
            "total": block["total"],
            "cells": [
                {"grade": grade, "count": counts.get(grade, 0)}
                for grade in columns
            ],
        })

    return {"columns": columns, "rows": rows}


def _broadsheet_student_summary(marks_per_subject, subjects, is_844, class_id):
    """Total points and mean grade for one learner on the broadsheet."""
    def subject_id(subject):
        return subject["id"] if isinstance(subject, dict) else subject.id

    def subject_category(subject):
        if isinstance(subject, dict):
            return subject.get("category") or ""
        return subject.category or ""

    if is_844:
        from .grading_844 import compute_844_aggregate

        points_rows = []
        for subject in subjects:
            mark_info = marks_per_subject.get(subject_id(subject), {}) or {}
            if mark_info.get("points") is not None:
                points_rows.append({
                    "points": mark_info["points"],
                    "category": subject_category(subject),
                })

        if not points_rows:
            return None, None

        summary = compute_844_aggregate(points_rows)
        return summary["total_points"], summary["mean_grade"]

    points_total = 0
    has_points = False
    numeric = []
    for subject in subjects:
        mark_info = marks_per_subject.get(subject_id(subject), {}) or {}
        pts = mark_info.get("points")
        if pts is not None:
            points_total += pts
            has_points = True
        raw = mark_info.get("marks")
        if raw in (None, "-", ""):
            continue
        try:
            numeric.append(float(raw))
        except (TypeError, ValueError):
            continue

    mean_grade = None
    if numeric:
        mean = int(round(sum(numeric) / len(numeric)))
        grade_info = resolve_grade(class_id, mean)
        mean_grade = grade_info.get("performance_level") if grade_info else None

    return (points_total if has_points else None), mean_grade


def _student_ranking_score(student, subjects, is_844):
    if is_844:
        from .grading_844 import compute_844_aggregate

        points_rows = []
        for subj in subjects:
            mark_info = student["marks"].get(subj["id"], {})
            if mark_info.get("points") is not None:
                points_rows.append({
                    "points": mark_info["points"],
                    "category": subj.get("category", ""),
                })

        if not points_rows:
            return 0, "—", 0

        summary = compute_844_aggregate(points_rows)
        return summary["total_points"], summary["mean_grade"], len(points_rows)

    numeric_marks = [
        mark_info["marks"]
        for mark_info in student["marks"].values()
        if mark_info.get("marks") not in (None, "-")
    ]

    if not numeric_marks:
        return 0, "—", 0

    total = sum(numeric_marks)
    mean = int(round(total / len(numeric_marks)))
    return total, mean, len(numeric_marks)


def compute_full_analysis(data):
    """
    Build rankings and grade-distribution analytics for the full analysis PDF.
    """
    subjects = data.get("subjects") or []
    students = data.get("students") or []
    grading_type = data.get("grading_type", "cbc")
    is_844 = grading_type == "844"

    empty = {
        "top_students": [],
        "top_per_subject": [],
        "best_boy": None,
        "best_girl": None,
        "subject_grade_breakdown": [],
        "subject_grade_table": {"columns": [], "rows": []},
        "overall_grade_analysis": [],
        "overall_grade_title": "Overall Grade Analysis",
        "overall_grade_subtitle": "",
        "grade_label": "Mean Grade" if is_844 else "Mean Score",
        "score_label": "Total Points" if is_844 else "Total Marks",
    }

    if not students or not subjects:
        return empty

    ranked_students = []
    for student in students:
        score, mean_value, subjects_count = _student_ranking_score(
            student, subjects, is_844
        )
        ranked_students.append({
            "id": student["id"],
            "name": student["full_name"],
            "admission_number": student.get("admission_number"),
            "gender": student.get("gender"),
            "score": score,
            "mean_value": mean_value,
            "subjects_count": subjects_count,
        })

    ranked_students.sort(
        key=lambda row: (row["score"], row["name"]),
        reverse=True,
    )

    top_students = []
    for position, row in enumerate(ranked_students[:5], start=1):
        top_students.append({**row, "position": position})

    best_boy = None
    best_girl = None
    for row in ranked_students:
        if best_boy is None and _is_male(row["gender"]):
            best_boy = row
        if best_girl is None and _is_female(row["gender"]):
            best_girl = row
        if best_boy and best_girl:
            break

    top_per_subject = []
    subject_grade_detail = {subj["id"]: {} for subj in subjects}
    overall_grades = {}

    for subj in subjects:
        best_entry = None

        for student in students:
            mark_info = student["marks"].get(subj["id"], {})
            marks = mark_info.get("marks")
            grade = mark_info.get("grade")

            if grade:
                subject_grade_detail[subj["id"]][grade] = (
                    subject_grade_detail[subj["id"]].get(grade, 0) + 1
                )
                if not is_844:
                    overall_grades[grade] = overall_grades.get(grade, 0) + 1

            if marks in (None, "-"):
                continue

            if best_entry is None or marks > best_entry["marks"]:
                best_entry = {
                    "student": student["full_name"],
                    "marks": int(float(marks)),
                    "grade": grade or "—",
                    "grade_css": _grade_css_class(grade),
                }

        if best_entry:
            top_per_subject.append({
                "subject": subj["name"],
                "student": best_entry["student"],
                "marks": best_entry["marks"],
                "grade": best_entry["grade"],
                "grade_css": best_entry["grade_css"],
            })

    subject_grade_breakdown = []
    for subj in subjects:
        grades = _sort_grade_counts(
            subject_grade_detail.get(subj["id"], {}),
            grading_type,
        )
        if grades:
            total = sum(item["count"] for item in grades)
            max_count = max(item["count"] for item in grades)
            for item in grades:
                item["percent"] = round((item["count"] / total) * 100, 1) if total else 0
                item["bar_width"] = round((item["count"] / max_count) * 100, 1) if max_count else 0

            subject_grade_breakdown.append({
                "subject": subj["name"],
                "total": total,
                "grades": grades,
            })

    subject_grade_table = _build_subject_grade_table(
        subject_grade_breakdown,
        grading_type,
    )

    if is_844:
        student_mean_grade_counts = {}
        for row in ranked_students:
            mean_grade = row["mean_value"]
            if mean_grade and mean_grade != "—":
                student_mean_grade_counts[mean_grade] = (
                    student_mean_grade_counts.get(mean_grade, 0) + 1
                )
        overall_grade_analysis = _sort_grade_counts(
            student_mean_grade_counts,
            grading_type,
        )
    else:
        overall_grade_analysis = _sort_grade_counts(overall_grades, grading_type)

    return {
        "top_students": top_students,
        "top_per_subject": top_per_subject,
        "best_boy": best_boy,
        "best_girl": best_girl,
        "subject_grade_breakdown": subject_grade_breakdown,
        "subject_grade_table": subject_grade_table,
        "overall_grade_analysis": overall_grade_analysis,
        "overall_grade_title": (
            "Overall Mean Grade Analysis"
            if is_844
            else "Overall Grade Analysis"
        ),
        "overall_grade_subtitle": (
            "Number of students by general mean grade"
            if is_844
            else "Grade distribution across all subject scores"
        ),
        "grade_label": "Mean Grade" if is_844 else "Mean Score",
        "score_label": "Total Points" if is_844 else "Total Marks",
    }