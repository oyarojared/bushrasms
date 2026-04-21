import os
from io import BytesIO
from secrets import token_hex

from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image
from werkzeug.utils import secure_filename

from .general_utils import allowed_file


def generate_excel_file(headers=[], fields=[], data=[]):
    wb = Workbook()
    ws = wb.active

    # ---- HEADER SECTION ---- #
    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    # ---- BODY SECTION ---- #
    row = 1
    for d in data:
        row += 1
        for col, field in enumerate(fields, start=1):
            value = getattr(d, field, "")

            # Format dates
            if hasattr(value, "strftime"):
                value = value.strftime("%Y-%m-%d")

            # ---- EXCEL INJECTION PROTECTION (ADDED) ---- #
            if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
                value = f"'{value}"
            # ------------------------------------------- #

            cell = ws.cell(row=row, column=col, value=value)

            # center content of specific cells.
            centered_fields = [
                "admission_number", "kcpe_year", "gender",
                "kcpe_marks", "staff_id", "id_no",
                "employer", "tsc_no"
            ]
            if field in centered_fields:
                cell.alignment = Alignment(horizontal="center")

    # AUTO-WIDTH
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(c.value)) for c in col if c.value)
        ws.column_dimensions[col_letter].width = max(15, min(max_len + 2, 50))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output



def preprocess_image(uploaded_img, size=(200, 200)):
    if not uploaded_img:
        return None

    filename = secure_filename(uploaded_img.filename)
    _, ext = os.path.splitext(filename)

    if not allowed_file(filename):
        raise ValueError("Unsupported image format")

    new_name = f"{token_hex(16)}{ext.lower()}"
    folder = os.path.join(current_app.root_path, "static/uploads/passports")
    os.makedirs(folder, exist_ok=True)

    img = Image.open(uploaded_img)
    img.thumbnail(size)
    img.save(os.path.join(folder, new_name), quality=95)

    return new_name
